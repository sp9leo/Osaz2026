"""
Usage:
  python create_template.py              → creates teachers_template.xlsx
  python create_template.py export       → reads teachers_template.xlsx → teachers.json
"""
import sys
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Potrebujem openpyxl: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).parent
TEMPLATE = BASE_DIR / "teachers_template.xlsx"
OUTPUT = BASE_DIR / "teachers.json"


def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Učitelji"

    headers = ["Ime (za mapo)", "Prikazno ime", "PIN"]
    ws.append(headers)

    style = ws["A1"]
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12

    examples = [
        ["Novak_Jan", "Novak Jan", "1001"],
        ["Horvat_Marjan", "Horvat Marjan", "1002"],
        ["Ahlin_Tina", "Ahlin Tina", "1003"],
    ]
    for row in examples:
        ws.append(row)

    ws2 = wb.create_sheet("Navodila")
    ws2["A1"] = "Navodila"
    ws2["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    ws2["A3"] = "Stolpec A - Ime (za mapo):"
    ws2["A4"] = "  Ime v formatu Priimek_Ime. To bo ime mape v uploads/."
    ws2["A5"] = "  Primer: Novak_Jan"
    ws2["A7"] = "Stolpec B - Prikazno ime:"
    ws2["A8"] = "  Lepo ime, ki se prikaže v aplikaciji."
    ws2["A9"] = "  Primer: Novak Jan"
    ws2["A11"] = "Stolpec C - PIN:"
    ws2["A12"] = "  4-mestna številka. Vsak učitelj ima svojo."
    ws2["A14"] = "Ko izpolnite tabelo, zaženite:"
    ws2["A15"] = "  python create_template.py export"
    ws2.column_dimensions["A"].width = 60

    wb.save(str(TEMPLATE))
    print(f"Ustvarjeno: {TEMPLATE}")


def export_json():
    if not TEMPLATE.exists():
        print(f"Najprej ustvarite predlogo: python create_template.py")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(TEMPLATE))
    ws = wb["Učitelji"]

    teachers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, display, pin = row[0], row[1], row[2]
        if not name or not pin:
            continue
        teachers.append({
            "name": str(name).strip(),
            "display": str(display).strip() if display else str(name).strip(),
            "pin": str(int(pin)).strip(),
        })

    config = {
        "admin_pin": "9999",
        "teachers": teachers,
    }

    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    print(f"Izvoženo: {OUTPUT} ({len(teachers)} učiteljev)")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "export":
        export_json()
    else:
        create_template()
